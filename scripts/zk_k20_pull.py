#!/usr/bin/env python3
"""
ZKTeco K20 Attendance Device – Full Data Pull
==============================================
Device info (from physical unit):
  Model        : K20
  Serial No.   : 3157153800040
  MAC          : 00:17:61:8B:CF:90
  Algorithm    : ZKFinger VX10.0
  Firmware     : Ver 4.0.1 (build 74)
  Vendor       : ZKTeco Inc.
  Manu. Time   : 2015-09-18

Usage:
  python zk_k20_pull.py --host <DEVICE_IP> --port 4370
  python zk_k20_pull.py --host <DEVICE_IP> --port 4370 --year 2025 --month 3
  python zk_k20_pull.py --host <DEVICE_IP> --port 4370 --force-udp
  python zk_k20_pull.py --host <DEVICE_IP> --port 4370 --device-info-only

Install deps:
  pip install pyzk openpyxl
"""
from __future__ import annotations

import argparse
import sys
import time
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("[ERROR] openpyxl not found. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    from zk import ZK
    from zk.exception import ZKNetworkError
except ImportError:
    print("[ERROR] pyzk not found. Run: pip install pyzk", file=sys.stderr)
    sys.exit(1)


# ─── DEVICE DEFAULTS (K20 specific) ──────────────────────────────────────────
DEVICE_SERIAL = "3157153800040"
DEVICE_MODEL  = "K20"
DEFAULT_PORT  = 4370
DEFAULT_PASS  = 0        # K20 has no password set
DEFAULT_TIMEOUT = 20     # older firmware – give it extra time


# ─── CLI ──────────────────────────────────────────────────────────────────────
def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=f"Pull attendance data from ZKTeco {DEVICE_MODEL} and export to XLSX"
    )
    p.add_argument("--host", default="103.245.195.202",
                   help="Static IP address of the K20 device (default: 103.245.195.202)")
    p.add_argument("--port", type=int, default=DEFAULT_PORT,
                   help=f"TCP port (default: {DEFAULT_PORT})")
    p.add_argument("--password", type=int, default=DEFAULT_PASS,
                   help="Device comm password (default: 0 = none)")
    p.add_argument("--timeout", type=int, default=DEFAULT_TIMEOUT,
                   help=f"Socket timeout in seconds (default: {DEFAULT_TIMEOUT})")
    p.add_argument("--year", type=int, default=datetime.now().year,
                   help="Filter year (default: current year)")
    p.add_argument("--month", type=int, choices=range(1, 13), metavar="1-12",
                   help="Filter month 1-12. Omit for all months in the year.")
    p.add_argument("--all-years", action="store_true",
                   help="Export ALL records regardless of year/month")
    p.add_argument("--force-udp", action="store_true",
                   help="Use UDP instead of TCP (try if TCP handshake fails)")
    p.add_argument("--device-info-only", action="store_true",
                   help="Print device info and exit without exporting")
    p.add_argument("--clear-after", action="store_true",
                   help="DANGER: clear attendance log on device after export")
    p.add_argument("--output", default=None,
                   help="Custom output XLSX path")
    p.add_argument("--retry", type=int, default=2,
                   help="Number of connection retries (default: 2)")
    return p


# ─── OUTPUT PATH ─────────────────────────────────────────────────────────────
def resolve_output(args) -> Path:
    if args.output:
        return Path(args.output).expanduser().resolve()
    script_dir = Path(__file__).resolve().parent
    result_dir = script_dir / "result"
    result_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if args.all_years:
        tag = "_all"
    elif args.month:
        tag = f"_{args.year}-{args.month:02d}"
    else:
        tag = f"_{args.year}_all-months"
    return result_dir / f"K20_attendance{tag}_{stamp}.xlsx"


# ─── CONNECT WITH RETRY ───────────────────────────────────────────────────────
def connect_with_retry(host, port, password, timeout, force_udp, retries=2):
    """
    K20 (firmware 4.0.1) can be slow to respond.
    We try TCP first; if that fails we optionally try UDP.
    """
    protocols = [(False, "TCP")]
    if force_udp:
        protocols = [(True, "UDP"), (False, "TCP")]

    last_err = None
    for udp_flag, label in protocols:
        for attempt in range(1, retries + 2):
            zk = ZK(
                host,
                port=port,
                timeout=timeout,
                password=password,
                force_udp=udp_flag,
                ommit_ping=True,   # essential for NAT / non-pingable devices
            )
            try:
                print(f"[INFO] Attempt {attempt} via {label} → {host}:{port}", flush=True)
                conn = zk.connect()
                print(f"[INFO] Connected via {label}", flush=True)
                return conn
            except ZKNetworkError as e:
                last_err = e
                print(f"[WARN] {label} attempt {attempt} failed: {e}", flush=True)
                if attempt <= retries:
                    time.sleep(2)
            except Exception as e:
                last_err = e
                print(f"[WARN] {label} attempt {attempt} error: {e}", flush=True)
                if attempt <= retries:
                    time.sleep(2)

    raise ZKNetworkError(f"All connection attempts failed. Last error: {last_err}")


# ─── DEVICE INFO ─────────────────────────────────────────────────────────────
def print_device_info(conn) -> dict:
    info = {}
    try:
        info["firmware"]   = conn.get_firmware_version()
    except Exception:
        info["firmware"]   = "N/A"
    try:
        info["serial"]     = conn.get_serialnumber()
    except Exception:
        info["serial"]     = DEVICE_SERIAL  # fallback from label
    try:
        info["platform"]   = conn.get_platform()
    except Exception:
        info["platform"]   = "N/A"
    try:
        info["device_name"]= conn.get_device_name()
    except Exception:
        info["device_name"]= DEVICE_MODEL
    try:
        info["work_code"]  = conn.get_workcode()
    except Exception:
        info["work_code"]  = "N/A"
    try:
        info["fingerprint_algorithm"] = conn.get_fp_version()
    except Exception:
        info["fingerprint_algorithm"] = "ZKFinger VX10.0"

    print("\n── Device Info ─────────────────────────────")
    for k, v in info.items():
        print(f"  {k:<28}: {v}")
    print("────────────────────────────────────────────\n")
    return info


# ─── STYLED XLSX HEADER ───────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

def write_header(ws, cols: list[str]):
    ws.append(cols)
    for cell in ws[1]:
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = HEADER_ALIGN
    ws.row_dimensions[1].height = 18


# ─── PUNCH / STATUS LABELS ────────────────────────────────────────────────────
PUNCH_MAP  = {0: "Check-In", 1: "Check-Out", 2: "Overtime-In", 3: "Overtime-Out",
              4: "Break-Out", 5: "Break-In", 255: "Other"}
STATUS_MAP = {0: "Fingerprint", 1: "PIN", 2: "Card", 3: "Finger+PIN",
              4: "Finger+Card", 5: "PIN+Card", 6: "Finger+PIN+Card",
              15: "Face", 200: "Duress-Finger", 201: "AntiPassback",
              202: "Interlock"}


# ─── MAIN EXPORT ─────────────────────────────────────────────────────────────
def export(args, output: Path):
    conn = connect_with_retry(
        args.host, args.port, args.password, args.timeout,
        args.force_udp, args.retry
    )

    try:
        conn.disable_device()

        # ── device info ──
        dev_info = print_device_info(conn)
        if args.device_info_only:
            return 0, 0, 0

        # ── fetch data ──
        print("[INFO] Fetching user list...", flush=True)
        users = conn.get_users() or []
        print(f"[INFO] {len(users)} users found", flush=True)

        print("[INFO] Fetching attendance logs (may take a while on K20)...", flush=True)
        logs = conn.get_attendance() or []
        print(f"[INFO] {len(logs)} raw log entries fetched", flush=True)

        # ── build user lookup ──
        by_uid  = {str(getattr(u, "uid",     "")).strip(): u for u in users}
        by_uid2 = {str(getattr(u, "user_id", "")).strip(): u for u in users}

        # ── filter logs ──
        rows = []
        skipped = 0
        for log in logs:
            ts = getattr(log, "timestamp", None)
            if not ts:
                skipped += 1
                continue
            if not args.all_years:
                if ts.year != args.year:
                    skipped += 1
                    continue
                if args.month is not None and ts.month != args.month:
                    skipped += 1
                    continue

            uid_str = str(getattr(log, "user_id", "")).strip()
            uid_raw = str(getattr(log, "uid",     "")).strip()
            user    = by_uid2.get(uid_str) or by_uid.get(uid_raw)

            punch_raw  = getattr(log, "punch",  0)
            status_raw = getattr(log, "status", 0)

            rows.append({
                "timestamp":     ts,
                "date":          ts.date().isoformat(),
                "time":          ts.strftime("%H:%M:%S"),
                "day_of_week":   ts.strftime("%A"),
                "employee_name": getattr(user, "name", "") if user else "",
                "device_user_id":uid_str,
                "uid":           uid_raw,
                "punch_code":    punch_raw,
                "punch_label":   PUNCH_MAP.get(punch_raw, str(punch_raw)),
                "status_code":   status_raw,
                "status_label":  STATUS_MAP.get(status_raw, str(status_raw)),
                "card":          getattr(user, "card", "") if user else "",
                "privilege":     getattr(user, "privilege", "") if user else "",
            })

        rows.sort(key=lambda r: r["timestamp"])
        print(f"[INFO] {len(rows)} records after filter  ({skipped} skipped)", flush=True)

        # ── WORKBOOK ──────────────────────────────────────────────────────────
        wb = Workbook()

        # Sheet 1 – Attendance Logs
        ws_att = wb.active
        ws_att.title = "Attendance Logs"
        ATT_COLS = [
            "Timestamp", "Date", "Day", "Time",
            "EmployeeName", "DeviceUserID", "UID",
            "PunchCode", "PunchLabel", "StatusCode", "StatusLabel",
            "Card", "Privilege",
        ]
        write_header(ws_att, ATT_COLS)
        for r in rows:
            ws_att.append([
                r["timestamp"].strftime("%Y-%m-%d %H:%M:%S"),
                r["date"],
                r["day_of_week"],
                r["time"],
                r["employee_name"],
                r["device_user_id"],
                r["uid"],
                r["punch_code"],
                r["punch_label"],
                r["status_code"],
                r["status_label"],
                r["card"],
                r["privilege"],
            ])
        ATT_WIDTHS = [22, 12, 12, 10, 28, 14, 8, 11, 14, 12, 16, 16, 10]
        for i, w in enumerate(ATT_WIDTHS, 1):
            ws_att.column_dimensions[
                ws_att.cell(1, i).column_letter
            ].width = w

        # Sheet 2 – User Master
        ws_usr = wb.create_sheet("User Master")
        USR_COLS = ["UID", "DeviceUserID", "EmployeeName", "Card", "Privilege", "GroupID"]
        write_header(ws_usr, USR_COLS)
        for u in sorted(users, key=lambda x: str(getattr(x, "user_id", "")).strip()):
            ws_usr.append([
                getattr(u, "uid",       ""),
                str(getattr(u, "user_id", "")).strip(),
                getattr(u, "name",      ""),
                getattr(u, "card",      ""),
                getattr(u, "privilege", ""),
                getattr(u, "group_id",  ""),
            ])
        USR_WIDTHS = {"A": 8, "B": 14, "C": 30, "D": 18, "E": 10, "F": 12}
        for col, w in USR_WIDTHS.items():
            ws_usr.column_dimensions[col].width = w

        # Sheet 3 – Device Info
        ws_dev = wb.create_sheet("Device Info")
        write_header(ws_dev, ["Field", "Value"])
        ws_dev.column_dimensions["A"].width = 28
        ws_dev.column_dimensions["B"].width = 36
        for k, v in dev_info.items():
            ws_dev.append([k, str(v)])
        ws_dev.append(["export_timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws_dev.append(["export_filter_year",  str(args.year) if not args.all_years else "ALL"])
        ws_dev.append(["export_filter_month", str(args.month) if args.month else "ALL"])

        # Sheet 4 – Daily Summary
        ws_sum = wb.create_sheet("Daily Summary")
        SUM_COLS = ["Date", "Day", "EmployeeName", "DeviceUserID",
                    "FirstIn", "LastOut", "TotalPunches"]
        write_header(ws_sum, SUM_COLS)
        ws_sum.column_dimensions["A"].width = 12
        ws_sum.column_dimensions["B"].width = 12
        ws_sum.column_dimensions["C"].width = 28
        ws_sum.column_dimensions["D"].width = 14
        ws_sum.column_dimensions["E"].width = 12
        ws_sum.column_dimensions["F"].width = 12
        ws_sum.column_dimensions["G"].width = 14

        # group by (date, device_user_id)
        summary: dict[tuple, dict] = {}
        for r in rows:
            key = (r["date"], r["device_user_id"])
            if key not in summary:
                summary[key] = {
                    "day":           r["day_of_week"],
                    "name":          r["employee_name"],
                    "times":         [],
                }
            summary[key]["times"].append(r["time"])

        for (date, uid_s), data in sorted(summary.items()):
            times = sorted(data["times"])
            ws_sum.append([
                date,
                data["day"],
                data["name"],
                uid_s,
                times[0],
                times[-1],
                len(times),
            ])

        # ── save ──
        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)

        mapped = sum(1 for r in rows if r["employee_name"])

        # ── optional clear ──
        if args.clear_after:
            print("[WARN] Clearing attendance log on device...", flush=True)
            conn.clear_attendance()
            print("[WARN] Attendance log cleared.", flush=True)

        return len(rows), len(users), mapped

    finally:
        try:
            conn.enable_device()
        except Exception:
            pass
        try:
            conn.disconnect()
        except Exception:
            pass


# ─── ENTRY POINT ─────────────────────────────────────────────────────────────
def main():
    args   = build_parser().parse_args()
    output = resolve_output(args)

    print(f"\n{'='*52}")
    print(f"  ZKTeco {DEVICE_MODEL} Data Pull")
    print(f"  Target : {args.host}:{args.port}")
    if not args.all_years:
        month_str = f"{args.month:02d}" if args.month else "all"
        print(f"  Filter : year={args.year}  month={month_str}")
    else:
        print(f"  Filter : ALL records")
    print(f"{'='*52}\n")

    try:
        rows, users, mapped = export(args, output)

        if args.device_info_only:
            return

        print(f"\n{'='*52}")
        print(f"  Output             : {output}")
        print(f"  Attendance rows    : {rows}")
        print(f"  Users on device    : {users}")
        print(f"  Rows with name     : {mapped}")
        print(f"  Rows without name  : {rows - mapped}")
        print(f"{'='*52}\n")

    except ZKNetworkError as e:
        print(f"\n[ERROR] Connection failed: {e}", file=sys.stderr)
        print("\nTroubleshooting for ZKTeco K20 (firmware 4.0.1):", file=sys.stderr)
        print("  1. Confirm the device IP is reachable:  ping <DEVICE_IP>", file=sys.stderr)
        print("  2. Confirm port is open:                nc -zv <DEVICE_IP> 4370", file=sys.stderr)
        print("  3. Try UDP mode:                        --force-udp", file=sys.stderr)
        print("  4. Try longer timeout:                  --timeout 30", file=sys.stderr)
        print("  5. Check device is on the same LAN / VPN segment", file=sys.stderr)
        sys.exit(1)

    except Exception as e:
        print(f"\n[ERROR] Unexpected: {e}", file=sys.stderr)
        import traceback; traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
